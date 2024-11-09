"""
Desafío 1: Gestor de notas con Tkinter y almacenamiento en archivo CSV

Descripción:
Crear una interfaz gráfica con tkinter que permita al usuario ingresar su nombre y una nota en un campo de texto,
y luego guardar estos datos en un archivo CSV.

El código debe:
- Verificar que los campos no estén vacíos.
- Validar que la nota sea un número.
- Almacenar los datos en un archivo CSV llamado 'notas.csv'.
- Mostrar un mensaje de éxito o error según sea necesario.
- Si el archivo CSV no existe, debe inicializarlo con encabezados.

Requisitos:
- Utilizar tkinter para la interfaz gráfica.
- Utilizar el módulo csv para manejar el archivo de notas.

"""
import openpyxl
import os
from openpyxl.utils import get_column_letter

def update_excel(file_path):
    # Comprobamos si el archivo tiene la extensión .xlsx
    if not file_path.endswith('.xlsx'):
        print("Error: El archivo debe tener la extensión '.xlsx'.")
        return

    # Si el archivo no existe, lo creamos y agregamos un encabezado
    if not os.path.exists(file_path):
        try:
            print(f"El archivo {file_path} no existe. Creando un nuevo archivo...")
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["Nombre", "Nota", "Asignatura"])
            wb.save(file_path)
            print("Nuevo archivo creado y encabezado agregado.")
        except Exception as e:
            print(f"Error al crear el archivo: {e}")
            return

    # Intentamos cargar el archivo existente
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except Exception as e:
        print(f"Error al abrir el archivo: {e}")
        return

    # Datos a agregar en la nueva fila
    new_row = ["Juan", "80", "Matemáticas"]
    sheet.append(new_row)

    # Guardar los cambios en el archivo
    try:
        wb.save(file_path)
        print("Los datos fueron agregados correctamente al archivo Excel.")
    except PermissionError:
        print("Error: No se puede guardar el archivo. Asegúrate de que esté cerrado.")
        return
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")
        return

    # Ajustar el tamaño de las columnas según el contenido
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Cerrar el archivo después de modificarlo
    wb.close()

# Llamada al script
file_path = "notas.xlsx"
update_excel(file_path)
